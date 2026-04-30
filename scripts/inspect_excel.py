"""
Step 1: Inspect the Excel file to understand its structure.
"""
import sys
import openpyxl
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r"C:/Users/rontf/EQ_Master_Database_temp.xlsx", read_only=True, data_only=True)
print("Sheet names:", wb.sheetnames)
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== Sheet: {sheet_name} ===")
    rows = list(ws.iter_rows(min_row=1, max_row=20, values_only=True))
    for i, row in enumerate(rows, 1):
        # Only print rows that have non-None content
        if any(c is not None for c in row):
            print(f"  Row {i}: {row}")
    print()

wb.close()
